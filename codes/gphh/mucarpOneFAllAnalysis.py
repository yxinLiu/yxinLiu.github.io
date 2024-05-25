# -*- coding: utf-8 -*-
#    gp for solving multi-vehicle uncapacitated ARP designed by yuxin
#    与静态相比，读的数据要改变，评价方式换成不确定环境及平均值，
#    多辆车并行，用一个list存储时刻表
#    车辆选择下一个任务的方式是：车-任务选前n个，然后任务-车确定一个。
#    Step1: compared with the baseline GPHH
#    One for all strategy

import operator
import math
import random

import numpy
import copy
import datetime
from deap import algorithms
from deap import base
from deap import creator
from deap import tools
from deap import gp
#from sympy import *
from deap.gp import Terminal, Primitive

from openpyxl import Workbook
from openpyxl import load_workbook
import os


from util.excelUtil import ExcelUtil

import sys
import readfile
import Path

runID = 0

class Vehicle:
    def __init__(self, time, vehicleID, vehicleLoad, vehiclePosition):
        self.time = time
        self.vehicleID = vehicleID
        self.vehicleLoad = vehicleLoad
        self.vehiclePosition = vehiclePosition
        #self.vehicleTour = vehicleTour

# Define new functions
def protectedDiv(left, right):
    try:
        return left / right
    except ZeroDivisionError:
        return 1

def max(left, right):
    if left >= right:
        return left
    else:
        return right

def min(left, right):
    if left <=right:
        return left
    else:
        return right

pset = gp.PrimitiveSet("MAIN", 10)
pset.addPrimitive(operator.add, 2)
pset.addPrimitive(operator.sub, 2)
pset.addPrimitive(operator.mul, 2)
pset.addPrimitive(protectedDiv, 2)
pset.addPrimitive(max, 2)
pset.addPrimitive(min, 2)
#pset.addPrimitive(math.sin, 1)
#pset.addPrimitive(math.atan2, 2)
pset.addEphemeralConstant("rand101", lambda: round(random.uniform(0, 1), 2))
# change the name of arguments from ARG0 to x
pset.renameArguments(ARG0='CFH') #Cost From Here (the current node) to the head node of the candidate task---1. 需要归一化吗？不需要
#pset.renameArguments(ARG1="CFR1") #Cost From the closest feasible alternative Route to the candidate task???---2。不理解
pset.renameArguments(ARG1="CR") #Cost to Refill (from the current node to the depot)
pset.renameArguments(ARG2="CTD") #Cost from the tail node of the candidate task To the Depot??---3。是一个candidate序列吗？
pset.renameArguments(ARG3="CTT1") #Cost from the tail of the candidate task To its closest unserved Task (the head)
pset.renameArguments(ARG4="DEM") #DEMand of the candidate task---4.怎么感觉是所有candidata task的累加和
pset.renameArguments(ARG5="DEM1") #DEMand of the closet unserved task to the candidata task
pset.renameArguments(ARG6="FRT") #Fraction of the Remaining (unserved) Tasks
#pset.renameArguments(ARG8="FUT") #Fraction of the Unassigned Tasks---5。在我这里和FRT一样啊
pset.renameArguments(ARG7="FULL") #FULLness of the vehicle (current load over capacity)
pset.renameArguments(ARG8="RQ") #Remaining Capacity of the vehicle
#pset.renameArguments(ARG11="RQ1") #Remaining Capacity for the closet alternative route
pset.renameArguments(ARG9="SC") #Serving Cost of the candidata task
#pset.renameArguments(ARG10="CFR1") #Cost From the closest feasible alternative Route to the candidate task

# creating fitness function and individual
# 适应度表现为base模块中的Fitness基类，个体类表现为一个gp.PrimitiveTree
creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMin)

#register some parameters，使用register将自定函数填充到工具箱base.Toolbox()当中，后可通过toolbox.name调用。
#过程中所需参数动态绑定
toolbox = base.Toolbox()
toolbox.register("expr", gp.genHalfAndHalf, pset=pset, min_=1, max_=2) #这里的1和2要check一下！！！！！！！！！
toolbox.register("individual", tools.initIterate, creator.Individual, toolbox.expr)
# list和toolbox.individual为tool.initRepeat的参数，剩余一个参数在下面使用的时候传入，即n=300。
toolbox.register("population", tools.initRepeat, list, toolbox.individual)
toolbox.register("compile", gp.compile, pset=pset)

def showPath(s, e, instRoute, path):
    if s == e:
        instRoute.append(e)
    else:
        if path[s][e] != s:
            showPath(s, path[s][e], instRoute, path)
        if path[path[s][e]][e] != path[s][e]:
            showPath(path[s][e], e, instRoute, path)
        else:
            instRoute.append(path[s][e])

def floyd(verticeNum, arcList):
    #除计算两点之间的最短路径dis外，还应记录最短路径的路线，记录在instRoute中
    dis=[None] * verticeNum
    path = [None] * verticeNum
    pathList = []
    for i in range(verticeNum):
        dis[i]=[None] * verticeNum
        path[i] = [None] * verticeNum
        for j in range(verticeNum):
            path[i][j] = i
            if i == j:
                dis[i][j] = 0.0
            else:
                dis[i][j] = 99999.0
    for i in range(len(arcList)):
        head = arcList[i].getHead()
        tail = arcList[i].getTail()
        if arcList[i].getConnect() == 1:
            dis[head][tail] = arcList[i].getDeadCost() #用估计的值来计算，并且除去实际不存在，其余情况不再重新计算最短路径
    # floyd algorithm
    for k in range(verticeNum):
        for i in range(verticeNum):
            for j in range(verticeNum):
                if dis[i][j] > dis[i][k] + dis[k][j]:
                    dis[i][j] = dis[i][k] + dis[k][j]
                    path[i][j] = k
    # recall the route
    for start in range(verticeNum):
        for end in range(verticeNum):
            instPath = Path.Path()
            instRoute = []
            instPath.setHead(start)
            instPath.setTail(end)
            showPath(start, end, instRoute, path)
            instRoute.append(end)
            instPath.setRoute(instRoute)
            #print instRoute
            pathList.append(instPath)
    return dis, pathList


def travel(start, end, dis, pathList, verticeNum, arcList, depot, passDepot): #start和end均为int类型
    #首先需要获取从start到end需要经过哪些边
    travelCost = 0.0
    if start == end:
        return travelCost, dis, pathList, passDepot
    else:
        count = start * verticeNum + end
        route = pathList[count].getRoute()
        index = -1
        for t in range(len(route)-1):
            head = route[t]
            tail = route[t+1]
            for i in range(len(arcList)):
                if arcList[i].getHead() == head and arcList[i].getTail() == tail:
                    index = i
                    break
            if arcList[index].getStoDeadCost() > 0:
                travelCost = travelCost + arcList[index].getStoDeadCost()
                if tail == depot:
                    passDepot = 1
            else:
                arcList[index].setConnect(0)
                if index % 2 == 0:
                    arcList[index + 1].setConnect(0)
                else:
                    arcList[index - 1].setConnect(0)
                dis, pathList = floyd(verticeNum, arcList)
                thisTravelCost, dis, pathList, passDepot = travel(head, end, dis, pathList, verticeNum, arcList, depot, passDepot)
                travelCost = travelCost + thisTravelCost
                break
        return travelCost, dis, pathList, passDepot


def calMaxCost(unEdgeList, load, vehiclePosition, depot, dis):
    maxCost = 0.0
    for i in range(len(unEdgeList)):
        thisEdgeCost = 0
        head = unEdgeList[i].getHead()
        if load >= unEdgeList[i].getDemand():
            thisEdgeCost = unEdgeList[i].getSerCost() + dis[vehiclePosition][head]
        else:
            thisEdgeCost = unEdgeList[i].getSerCost() + dis[vehiclePosition][depot] + dis[depot][head]
        if(thisEdgeCost > maxCost):
            maxCost = thisEdgeCost
    return maxCost

def calMaxDepotCost(unEdgeList, dis, depot):
    maxDepotCost = 0.0
    for i in range(len(unEdgeList)):
        thisDepotCost = 0.0
        tail = unEdgeList[i].getTail()
        thisDepotCost = dis[tail][depot]
        if thisDepotCost > maxDepotCost:
            maxDepotCost = thisDepotCost
    return maxDepotCost


def evalRules(individual, sampleNum):
    # compile: Transform the tree expression in a callable function
    # 这里将individual的树形表现形式转换为可执行的形式
    func = toolbox.compile(expr=individual)
    # 如果把读文件放在这里，那么每评价一个个体，都需要读一次文件
    rf = readfile.graph()
    verticeNum, numReq, arcList, capacity, depot, vehiculous = rf.read()
    unEdgeTotalNum = len(arcList) / 2.0
    distance, firstPathList = floyd(verticeNum, arcList)
    # begin the evaluation
    # 5次取平均值
    totalTourCost = 0.0
    for batch in range(sampleNum):
        tourCost = 0.0
        # multiple vehicles
        timeList = []
        for v in range(vehiculous):
            vehicle = Vehicle(0.0, v, capacity, depot)
            timeList.append(vehicle)
        dis = distance
        pathList = copy.deepcopy(firstPathList)
        # 为第 #batch# 个不确定实例随机生成deadheading costs和demands
        numpy.random.seed(50)
        for i in range(0, len(arcList), 2):
            mu = arcList[i].getDeadCost()  # 期望
            sigma = 0.2 * mu  # 标准差
            num = 1  # 个数为1
            stoDeadCost = numpy.random.normal(mu, sigma, num)
            arcList[i].setStoDeadCost(round(stoDeadCost[0], 2))
            arcList[i+1].setStoDeadCost(round(stoDeadCost[0], 2))
            stoDemand = 0.0
            if stoDeadCost > 0:
                mu = arcList[i].getDemand()   # 期望
                sigma = 0.2 * mu  # 标准差
                num = 1
                stoD = numpy.random.normal(mu, sigma, num)
                stoDemand = round(stoD[0], 2)
            if stoDemand < 0:
                stoDemand = 0.0
            arcList[i].setStoDemand(stoDemand)
            arcList[i+1].setStoDemand(stoDemand)
            arcList[i].setConnect(1)
            arcList[i+1].setConnect(1)
            arcList[i].setRDF(1.0)#remaining demand fraction
            arcList[i+1].setRDF(1.0)
        unEdgeList = copy.deepcopy(arcList)
        #实例在线确定后，开始演化路线
        while len(unEdgeList) > 0:
            #maxCost = calMaxCost(unEdgeList, timeList[0].vehicleLoad, timeList[0].vehilePosition, depot, dis)
            #maxDepotCost = calMaxDepotCost(unEdgeList, dis, depot)
            # 不明白这里下面为什么要减1再除以2
            currentUnEdgeNum = len(unEdgeList) / 2.0
            candidateTask = []
            for i in range(len(unEdgeList)):
                if unEdgeList[i].getDemand() <= timeList[0].vehicleLoad: # 有filter
                    head = unEdgeList[i].getHead()
                    tail = unEdgeList[i].getTail()
                    currentCFH = dis[timeList[0].vehiclePosition][head]
                    currentCR = dis[timeList[0].vehiclePosition][depot]
                    currentCTD = dis[tail][depot]
                    minCTT1 =sys.float_info.max #要排除反方向的
                    CUTindex = -1 #the index of the closet unserved task
                    if i % 2 == 0:
                        reviseI = i + 1
                    else:
                        reviseI = i - 1
                    for remaining in range(len(unEdgeList)):
                        if remaining != i and remaining != reviseI:
                            if dis[tail][unEdgeList[remaining].getHead()] < minCTT1:
                                minCTT1 = dis[tail][unEdgeList[remaining].getHead()]
                                CUTindex = remaining
                    if CUTindex == -1:
                        currentCTT1 = 0
                        currentDEM1 = 0
                    else:
                        currentCTT1 = minCTT1
                        currentDEM1 = unEdgeList[CUTindex].getDemand()
                    currentDEM = unEdgeList[i].getDemand()
                    currentFRT = (unEdgeTotalNum - currentUnEdgeNum) /unEdgeTotalNum
                    currentFULL = (capacity - timeList[0].vehicleLoad) / capacity
                    currentRQ = timeList[0].vehicleLoad
                    currentSC = unEdgeList[i].getSerCost()
                    #currentCFR1 = dis[timeList[1].vehiclePosition][head]
                    #for ve in range(2, len(timeList)):
                        #if dis[timeList[ve].vehiclePosition][head] < currentCFR1:
                            #currentCFR1 = dis[timeList[ve].vehiclePosition][head]
                    try:
                        hv = func(currentCFH,currentCR,currentCTD,currentCTT1,currentDEM,currentDEM1,currentFRT,currentFULL,currentRQ,currentSC)
                    except:
                        print(individual)
                        print(hv)
                    if math.isinf(hv):
                        hv = 1
                    elif math.isnan(hv):
                        hv = 0
                    else:
                        hv = hv
                    unEdgeList[i].setHValue(hv)
                    # 根据 "车-任务规则" 确定一组待服务的任务，如有n辆车，则选出hv值最小的n个任务，将任务编号存入candidateTask列表中
                    candidateTask.insert(0, i)
                    for x in range(len(candidateTask)-1, -1, -1):
                        if unEdgeList[i].getHValue() >= unEdgeList[candidateTask[x]].getHValue():
                            candidateTask.insert(x+1, i)
                            del candidateTask[0]
                            break
                    if len(candidateTask) > vehiculous:
                        del candidateTask[-1]
            # 车1不能满足剩余的任务
            # 指定ID为0的车负责所有的failure tasks
            if len(candidateTask) == 0:
                # 车辆1要返回depot
                passDepot = 0
                thisTravelCost, dis, pathList, passDepot = travel(timeList[0].vehiclePosition, depot, dis, pathList,
                                                                  verticeNum, arcList, depot, passDepot)
                tourCost = tourCost + thisTravelCost
                timeList[0].time = timeList[0].time + thisTravelCost
                timeList[0].vehiclePosition = depot
                timeList[0].vehicleLoad = capacity
                if timeList[0].vehicleID == 0: #如果ID为0，那么返回后继续等待分配任务，否则直接结束当天的任务
                    for i in range(len(timeList) - 1, -1, -1):
                        if timeList[i].time <= timeList[0].time:
                            timeList.insert(i + 1, timeList[0])
                            break
                del timeList[0]
            else:
                # 如果候选任务集不为空，那么就根据 "任务-车规则" 确定一个唯一服务的任务
                # candidateTask集合中的任务都是车1能够满足的
                chosenTaskID = candidateTask[0]

                # First of all, travel to the head of the chosen task, input the current position of the vehicle, and the head of the chosen task,
                #    return the traveling cost of this process
                # if pass the depot on the way, its capacity is refilled.
                passDepot = 0
                thisTravelCost, dis, pathList, passDepot = travel(timeList[0].vehiclePosition, unEdgeList[chosenTaskID].getHead(), dis, pathList, verticeNum, arcList, depot, passDepot)
                if passDepot == 1:
                    timeList[0].vehicleLoad = capacity
                tourCost = tourCost + thisTravelCost
                timeList[0].time = timeList[0].time + thisTravelCost
                timeList[0].vehiclePosition = unEdgeList[chosenTaskID].getHead()
                # 到达该任务的起点后，首先可以获知该任务是否可以通行。
                if unEdgeList[chosenTaskID].getStoDeadCost() > 0: # 可以通行
                    if timeList[0].vehicleLoad >= unEdgeList[chosenTaskID].getRDF() * unEdgeList[chosenTaskID].getStoDemand():
                        # 可以正常服务完选中的task
                        thisServeCost = unEdgeList[chosenTaskID].getRDF() * unEdgeList[chosenTaskID].getSerCost() + (1.0 - unEdgeList[chosenTaskID].getRDF()) * unEdgeList[chosenTaskID].getStoDeadCost()
                        tourCost = tourCost + thisServeCost
                        timeList[0].vehicleLoad = timeList[0].vehicleLoad - unEdgeList[chosenTaskID].getRDF() * unEdgeList[chosenTaskID].getStoDemand()
                        timeList[0].vehiclePosition = unEdgeList[chosenTaskID].getTail()
                        if timeList[0].vehiclePosition == depot:
                            timeList[0].vehicleLoad = capacity
                        timeList[0].time = timeList[0].time + thisServeCost
                        unEdgeList[chosenTaskID].setRDF(0.0)
                        del unEdgeList[chosenTaskID]
                        if chosenTaskID % 2 == 0:
                            del unEdgeList[chosenTaskID]
                        else:
                            del unEdgeList[chosenTaskID - 1]
                        # 删除timeList的首位元素，然后将该车辆新的执行时刻加入到timeList中
                        for i in range(len(timeList) - 1, -1, -1):
                            if timeList[i].time <= timeList[0].time:
                                timeList.insert(i + 1, timeList[0])
                                break
                        del timeList[0]
                    else:
                        # route failure, back to the depot to refill, and then came back to complete the service--traditional
                        # route failure，如果ID为0，那么返回depot后，等待重新分配任务，否则返回depot后，结束当天任务
                        # 如果是ID为0的车发生了任务失败，有两种方式，一种是从哪里跌倒，从哪里爬起来；另外一种是可以排个序
                        # 先实现的另外一种
                        ratio = timeList[0].vehicleLoad / unEdgeList[chosenTaskID].getStoDemand() # vehicleLoad是车辆的剩余装载能力
                        thisServeCost = ratio * unEdgeList[chosenTaskID].getSerCost() + (1.0 - ratio) * unEdgeList[chosenTaskID].getStoDeadCost()
                        tourCost = tourCost + thisServeCost
                        timeList[0].time = timeList[0].time + thisServeCost
                        timeList[0].vehiclePosition = unEdgeList[chosenTaskID].getTail()#发现错误20210221
                        if timeList[0].vehiclePosition == depot:
                            timeList[0].vehicleLoad = capacity
                        leftRatio = unEdgeList[chosenTaskID].getRDF() - ratio
                        # back to the depot
                        passDepot = 0
                        thisTravelCost, dis, pathList, passDepot = travel(timeList[0].vehiclePosition, depot, dis, pathList, verticeNum, arcList, depot, passDepot)
                        tourCost = tourCost + thisTravelCost
                        timeList[0].time = timeList[0].time + thisTravelCost
                        timeList[0].vehiclePosition = depot
                        timeList[0].vehicleLoad = capacity
                        # 把这个服务一半的任务再放回候选集
                        unEdgeList[chosenTaskID].setRDF(leftRatio)
                        # 同样修改反方向的任务上的相关信息
                        if chosenTaskID % 2 == 0:
                            unEdgeList[chosenTaskID + 1].setRDF(leftRatio)
                        else:
                            unEdgeList[chosenTaskID - 1].setRDF(leftRatio)
                        # 如果ID为0，那么返回后继续等待分配任务，否则直接结束当天的任务
                        if timeList[0].vehicleID == 0:
                            for i in range(len(timeList) - 1, -1, -1):
                                if timeList[i].time <= timeList[0].time:
                                    timeList.insert(i + 1, timeList[0])
                                    break
                        del timeList[0]
                else:
                    #不可以通行，那么删掉该任务，重新为车辆分配下一个任务
                    #更新全局信息
                    h = unEdgeList[chosenTaskID].getHead()
                    t = unEdgeList[chosenTaskID].getTail()
                    for i in range(len(arcList)):
                        if arcList[i].getHead() == h and arcList[i].getTail() == t:
                            ID = i
                            break
                    arcList[ID].setConnect(0)
                    if ID % 2 == 0:
                        arcList[ID + 1].setConnect(0)
                    else:
                        arcList[ID - 1].setConnect(0)
                    dis, pathList = floyd(verticeNum, arcList)
                    del unEdgeList[chosenTaskID]
                    if chosenTaskID % 2 == 0:
                        del unEdgeList[chosenTaskID]
                    else:
                        del unEdgeList[chosenTaskID - 1]
                    # 删除timeList的首位元素，然后将该车辆新的执行时刻加入到timeList中
                    for i in range(len(timeList) - 1, -1, -1):
                        if timeList[i].time <= timeList[0].time:
                            timeList.insert(i + 1, timeList[0])
                            break
                    del timeList[0]
        #最后所有车辆还需返回depot
        for v in range(len(timeList)):
            passDepot = 0
            backC, dis, pathList, passDepot = travel(timeList[v].vehiclePosition, depot, dis, pathList, verticeNum, arcList, depot, passDepot)
            tourCost = tourCost + backC
        totalTourCost = totalTourCost + tourCost
    avgCost = totalTourCost / sampleNum
    #注意这个逗号，即使是单变量优化问题，也需要返回tuple（元组类型）
    return avgCost,

toolbox.register("evaluate", evalRules, sampleNum=5)
toolbox.register("select", tools.selTournament, tournsize=7)
# 将gp.cxOnePoint函数命名为"mate"，放入工具箱中，执行时，调用工具箱的"mate"，便会直接调用gp.cxOnePoint，进行交叉
toolbox.register("mate", gp.cxOnePoint)
toolbox.register("expr_mut", gp.genFull, min_=2, max_=4)
toolbox.register("mutate", gp.mutUniform, expr=toolbox.expr_mut, pset=pset)

toolbox.decorate("mate", gp.staticLimit(key=operator.attrgetter("height"), max_value=8))
toolbox.decorate("mutate", gp.staticLimit(key=operator.attrgetter("height"), max_value=8))

def gpSimple(population, toolbox, cxpb, mutpb, ngen, stats=None,
             halloffame=None, verbose=__debug__):
    """This algorithm reproduce the simplest evolutionary algorithm as
    presented in chapter 7 of [Back2000]_.

    :param population: A list of individuals.
    :param toolbox: A :class:`~deap.base.Toolbox` that contains the evolution
                    operators.
    :param cxpb: The probability of mating two individuals.
    :param mutpb: The probability of mutating an individual.
    :param ngen: The number of generation.
    :param stats: A :class:`~deap.tools.Statistics` object that is updated
                  inplace, optional.
    :param halloffame: A :class:`~deap.tools.HallOfFame` object that will
                       contain the best individuals, optional.
    :param verbose: Whether or not to log the statistics.
    :returns: The final population
    :returns: A class:`~deap.tools.Logbook` with the statistics of the
              evolution

    The algorithm takes in a population and evolves it in place using the
    :meth:`varAnd` method. It returns the optimized population and a
    :class:`~deap.tools.Logbook` with the statistics of the evolution. The
    logbook will contain the generation number, the number of evaluations for
    each generation and the statistics if a :class:`~deap.tools.Statistics` is
    given as argument. The *cxpb* and *mutpb* arguments are passed to the
    :func:`varAnd` function. The pseudocode goes as follow ::

        evaluate(population)
        for g in range(ngen):
            population = select(population, len(population))
            offspring = varAnd(population, toolbox, cxpb, mutpb)
            evaluate(offspring)
            population = offspring

    As stated in the pseudocode above, the algorithm goes as follow. First, it
    evaluates the individuals with an invalid fitness. Second, it enters the
    generational loop where the selection procedure is applied to entirely
    replace the parental population. The 1:1 replacement ratio of this
    algorithm **requires** the selection procedure to be stochastic and to
    select multiple times the same individual, for example,
    :func:`~deap.tools.selTournament` and :func:`~deap.tools.selRoulette`.
    Third, it applies the :func:`varAnd` function to produce the next
    generation population. Fourth, it evaluates the new individuals and
    compute the statistics on this population. Finally, when *ngen*
    generations are done, the algorithm returns a tuple with the final
    population and a :class:`~deap.tools.Logbook` of the evolution.

    .. note::

        Using a non-stochastic selection method will result in no selection as
        the operator selects *n* individuals from a pool of *n*.

    This function expects the :meth:`toolbox.mate`, :meth:`toolbox.mutate`,
    :meth:`toolbox.select` and :meth:`toolbox.evaluate` aliases to be
    registered in the toolbox.

    .. [Back2000] Back, Fogel and Michalewicz, "Evolutionary Computation 1 :
       Basic Algorithms and Operators", 2000.
    """
    #st = datetime.datetime.now()
    starttime = datetime.datetime.now()

    logbook = tools.Logbook()
    logbook.header = ['gen', 'nevals'] + (stats.fields if stats else [])

    # Evaluate the individuals with an invalid fitness
    invalid_ind = [ind for ind in population if not ind.fitness.valid]
    fitnesses = toolbox.map(toolbox.evaluate, invalid_ind)
    for ind, fit in zip(invalid_ind, fitnesses):
        ind.fitness.values = fit

    if halloffame is not None:
        halloffame.update(population)

    record = stats.compile(population) if stats else {}
    logbook.record(gen=0, nevals=len(invalid_ind), **record)

    #创建文件，将结果写入文件中
    #fileName = "/Users/liuyx/PycharmProjects/ga/output%s.xlsx" % starttime
    fileName = "gdb13%s.xlsx"%starttime.strftime('%Y%m%d%H%M%S')
    sheetName = "OneFAll"
    excelUtil = ExcelUtil(fileName, sheetName)
    excelWb = excelUtil.createExcelSheet()

    aStr = ''
    if verbose:
        #aStr = logbook.stream
        #print aStr
        print(logbook.stream)

    # 打印最优个体
    print(halloffame.items[0])
    print(halloffame.keys[0]) #这是中间某一代最好的，不一定是最后一代中最好的。
    #第0代中最优个体在testing中的表现
    #avgTest = evalRules(halloffame.items[0], 500)
    #print "avgTest:", avgTest[0]
    endtime = datetime.datetime.now()
    time = endtime - starttime
    print(time)

    bStr = '%s'%halloffame.items[0]
    cStr = '%s'%halloffame.keys[0]


    excelUtil.wirteDBToExcelByWb(aStr, bStr, cStr, time)


    # Begin the generational process
    for gen in range(1, ngen + 1):
        # Select the next generation individuals
        offspring = toolbox.select(population, len(population))

        # Vary the pool of individuals
        offspring = algorithms.varAnd(offspring, toolbox, cxpb, mutpb)

        # Evaluate the individuals with an invalid fitness
        #invalid_ind = [ind for ind in offspring if not ind.fitness.valid]
        invalid_ind = [ind for ind in offspring]#changed by Yuxin because each individual should be evaluated again
        fitnesses = toolbox.map(toolbox.evaluate, invalid_ind)
        for ind, fit in zip(invalid_ind, fitnesses):
            ind.fitness.values = fit

        # Update the hall of fame with the generated individuals
        halloffame = tools.HallOfFame(1)
        if halloffame is not None:
            halloffame.update(offspring) #similar的情况，不影响结果，因为还是那个individual

        # Replace the current population by the offspring
        population[:] = offspring

        # Append the current generation statistics to the logbook
        record = stats.compile(population) if stats else {}
        logbook.record(gen=gen, nevals=len(invalid_ind), **record)
        if verbose:
            print(logbook.stream)
        # 打印最优个体
        print(halloffame.items[0])
        print(halloffame.keys[0])

        endtime = datetime.datetime.now()
        time = endtime - starttime
        print(time)

        aStr = '%s' % logbook.stream
        bStr = '%s' % halloffame.items[0]
        cStr = '%s' % halloffame.keys[0]

        excelUtil.wirteDBToExcelByWb(aStr, bStr, cStr, time)

        # test
        if gen == ngen:
            avgTest = evalRules(halloffame.items[0], 1)
            print("avgTest:", avgTest)
            excelUtil.wirteDBToExcelByWb('%s'%avgTest, '', '', '')

    excelUtil.saveWbObjToExcel(excelWb)
    return population, logbook

def listExcelsInDir(self, srcDir):
    filterStr = '.xlsx'
    srcExcelList = []
    for root, dirs, files in os.walk(srcDir, topdown=False):
        for name in files:
            if name.endswith(filterStr):
                excelFileName = os.path.join(root, name)
                srcExcelList.append(excelFileName)
    return srcExcelList

def main():
    CFH = Terminal("CFH",str ,object)
    CTD = Terminal("CTD",str ,object)
    DEM1 = Terminal("DEM1",str ,object)
    FULL = Terminal("FULL",str ,object)
    CR = Terminal("CR",str ,object)
    DEM= Terminal("DEM",str ,object)
    SC = Terminal("SC", str ,object)
    FRT = Terminal("FRT", str ,object)
    CTT1 = Terminal("CTT1", str, object)
    RQ = Terminal("RQ", str, object)
    terminal068 = Terminal(0.68,0,0)
    terminal011 = Terminal(0.11,0,0)
    terminal044 = Terminal(0.44, 0, 0)

    max = Primitive("max", [object, object], object)
    min = Primitive("min", [object, object], object)
    mul = Primitive("mul", [object, object], object)
    sub = Primitive("sub", [object, object], object)
    add = Primitive("add", [object, object], object)
    protectedDiv = Primitive("protectedDiv", [object, object], object)

    #list=[protectedDiv,CFH, CR]
    #pt = gp.PrimitiveTree(list)

    #list = [subP,mulP,addP,minP,addP,FULLT, CFHT, DEM1T, addP,divP,CFHT, terminal026, mulP,mulP,terminal085, FULLT, CTDT, addP,minP,SCT, FRTT, minP,divP,addP,divP,CFHT, FRTT, maxP,RQT, CTT1T, addP,minP,CTDT, RQT, mulP,RQT, CTDT, CFHT, addP,subP,mulP,CRT, FRTT, addP,minP,CFHT, mulP,mulP,CFHT, RQT, minP,CFHT, DEMT, divP,subP,addP,FULLT, CFHT, mulP,CTDT, DEMT, maxP,mulP,CRT, FRTT, divP,CTDT, CTDT, minP,terminal018, CRT]
    #list=[addP,divP,SCT, CFHT, mulP,maxP,CFHT, maxP,subP,subP,divP,CRT, SCT, subP,SCT, divP,FULLT, RQT, FULLT, subP,minP,CRT, CFHT, SCT, subP,divP,maxP,FULLT, CTT1T, FRTT, minP,maxP,FULLT, FULLT, subP,CTT1T, FRTT]
    #list实例13只有一辆车返回，其余无失败=[addT,CFHT, divP,divP,mulT,maxT,FULLT, DEMT, maxT,FULLT, FULLT, maxT,minT,divP,addT,CRT, CFHT, DEM1T, maxT,divP,addT,SCT, CRT, SCT, DEM1T, addT,CRT, DEM1T, minT,mulT,maxT,addT,addT,minT,terminal054, CTT1T, FULLT, subT,mulT,SCT, CTDT, subT,CTDT, DEM1T, addT,SCT, CRT, subT,CRT, CTDT, mulT,subT,FULLT, CFHT, minT,subT,CRT, CTDT, terminal071]
    #list648.29=[subT,CFHT, mulT,minT,maxT,maxT,SCT, CTDT, DEM1T, minT,divP,divP,CRT, DEM1T, DEM1T, DEM1T, divP,minT,minT,SCT, divP,DEM1T, SCT, maxT,DEM1T, CTT1T, addT,CRT, minT,minT,minT,CFHT, divP,DEM1T, SCT, maxT,addT,FULLT, DEMT, SCT, maxT,divP,subT,CTDT, CTT1T, terminal084, mulT,divP,CRT, DEMT, addT,CTDT, SCT]
    #list581.24=[addT,CRT, maxT,maxT,subT,mulT,CRT, FRTT, divP,addT,SCT, FRTT, DEM1T, mulT,mulT,FULLT, divP,mulT,CFHT, RQT, addT,DEM1T, FRTT, subT,CFHT, CRT, maxT,addT,addT,mulT,CFHT, maxT,FULLT, CTDT, CFHT, mulT,addT,subT,subT,FULLT, CTT1T, subT,CTDT, DEM1T, SCT, FULLT, mulT,subT,CFHT, FRTT, minT,mulT,maxT,maxT,DEM1T, CRT, CFHT, subT,divP,CTDT, DEM1T, subT,FRTT, SCT, addT,addT,addT,DEMT, DEMT, maxT,FRTT, CFHT, minT,maxT,DEMT, SCT, addT,SCT, FRTT]
    #list648.29=[addT,minT,divT,minT,mulT,FULLT, CTT1T, divT,DEM1T, SCT, addT,DEM1T, addT,SCT, divT,RQT, maxT,subT,DEM1T, SCT, SCT, maxT,addT,FULLT, addT,addT,FULLT, CFHT, divT,DEM1T, CFHT, minT,FRTT, DEM1T, divT,addT,FULLT, addT,maxT,addT,FULLT, CFHT, FULLT, CFHT, divT,FRTT, addT,mulT,subT,minT,CFHT, addT,CTDT, CFHT, minT,addT,SCT, RQT, subT,CTDT, DEM1T, minT,divT,maxT,FULLT, FULLT, mulT,SCT, CTDT, divT,maxT,FULLT, SCT, CTDT, addT,DEM1T, addT,SCT, DEM1T]
    #list580.60=[addT,addT,minT,addT,maxT,CFHT, DEMT, divT,DEMT, CTDT, divT,maxT,maxT,SCT, CRT, DEMT, addT,CTDT, CRT, minT,CFHT, CFHT, minT,divT,maxT,maxT,SCT, CRT, CFHT, maxT,addT,subT,addT,subT,CTDT, SCT, SCT, SCT, minT,DEMT, terminal031, addT,maxT,CFHT, DEMT, divT,DEMT, CTDT, minT,CFHT, maxT,SCT, CRT]
    destDir = 'E:\\PycharmProjects\\ga\\results\\gdb8\\OneFAll'
    destList = listExcelsInDir(None, destDir)
    srcArr = []
    # 创建文件，将结果写入文件中
    # fileName = "/Users/liuyx/PycharmProjects/ga/output%s.xlsx" % starttime
    fileName = "gdb8_OneFAll.xlsx"
    sheetName = "OneFAll"
    excelUtil = ExcelUtil(fileName, sheetName)
    excelWb = excelUtil.createExcelSheet()
    column=0

    for excelFileName in destList:
        column=column+1
        wb = load_workbook(excelFileName)
        sheets = wb.sheetnames
        # 读取第一个sheet页
        ws = wb[sheets[0]]
        for i in range(1, 53):
            heuFun=ws.cell(i,1).value
            heuFun2=heuFun.replace('(',',').replace(')','')
            list=heuFun2.split(',')
            listFun = []
            for j in range (len(list)):
                if list[j].strip() == 'max':
                    listFun.append(max)
                elif list[j].strip() == 'min':
                    listFun.append(min)
                elif list[j].strip() == 'mul':
                    listFun.append(mul)
                elif list[j].strip() == 'sub':
                    listFun.append(sub)
                elif list[j].strip() == 'add':
                    listFun.append(add)
                elif list[j].strip() == 'protectedDiv':
                    listFun.append(protectedDiv)
                elif list[j].strip() == 'CFH':
                    listFun.append(CFH)
                elif list[j].strip() == 'CTD':
                    listFun.append(CTD)
                elif list[j].strip() == 'DEM1':
                    listFun.append(DEM1)
                elif list[j].strip() == 'FULL':
                    listFun.append(FULL)
                elif list[j].strip() == 'CR':
                    listFun.append(CR)
                elif list[j].strip() == 'DEM':
                    listFun.append(DEM)
                elif list[j].strip() == 'SC':
                    listFun.append(SC)
                elif list[j].strip() == 'FRT':
                    listFun.append(FRT)
                elif list[j].strip() == 'CTT1':
                    listFun.append(CTT1)
                elif list[j].strip() == 'RQ':
                    listFun.append(RQ)
                else:
                    listFun.append(Terminal(float(list[j].strip()), 0, 0))

            pt = gp.PrimitiveTree(listFun)
            avgTest = evalRules(pt, 500)
            print("avgTest:", avgTest)
            excelUtil.wirteDBToExcelByRowAndColumn('%s' % avgTest, i, column)
            srcArr.append(avgTest)
            #print("avgTest:", avgTest)
        wb.close()
    excelUtil.saveWbObjToExcel(excelWb)
    #list=[protectedDiv,CFH, CR]
    #pt = gp.PrimitiveTree(list)
    #avgTest = evalRules(pt, 500)
    #print("avgTest:", avgTest)


if __name__ == "__main__":
    for run in range(1):
        main()
